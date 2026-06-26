"""Exportadores de reportes a Excel (openpyxl) y PDF (reportlab)."""

import io

from django.http import HttpResponse

VERDE = "0F766E"


def exportar_excel(titulo, columnas, filas, filename="reporte.xlsx"):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte"

    ws.append([titulo])
    ws["A1"].font = Font(bold=True, size=14)

    ws.append(columnas)
    header_fill = PatternFill("solid", fgColor=VERDE)
    for cell in ws[2]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill

    for fila in filas:
        ws.append(fila)

    for i, _ in enumerate(columnas, start=1):
        ws.column_dimensions[chr(64 + i)].width = 18

    buff = io.BytesIO()
    wb.save(buff)
    buff.seek(0)
    resp = HttpResponse(
        buff.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    resp["Content-Disposition"] = f'attachment; filename="{filename}"'
    return resp


def exportar_pdf(titulo, columnas, filas, filename="reporte.pdf"):
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import landscape, letter
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

    buff = io.BytesIO()
    doc = SimpleDocTemplate(buff, pagesize=landscape(letter), title=titulo)
    styles = getSampleStyleSheet()
    story = [Paragraph(titulo, styles["Title"]), Spacer(1, 12)]

    data = [columnas] + [[str(c) for c in fila] for fila in filas]
    tabla = Table(data, repeatRows=1)
    tabla.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0F766E")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#DFE7EF")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F4F7F6")]),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
    ]))
    story.append(tabla)
    doc.build(story)

    buff.seek(0)
    resp = HttpResponse(buff.getvalue(), content_type="application/pdf")
    resp["Content-Disposition"] = f'attachment; filename="{filename}"'
    return resp
